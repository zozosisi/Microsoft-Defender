"""Verify Excel report data matches CSV baseline."""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
from openpyxl import load_workbook

csv_df = pd.read_csv(r'..\incidents\analysis\user_investigation_summary.csv')
wb = load_workbook(r'..\incidents\analysis\investigation_report.xlsx')
ws = wb['User Investigation']

print("=== FULL COMPARISON (CSV vs Excel) ===")
mismatches = 0
for idx, (_, row) in enumerate(csv_df.iterrows()):
    excel_row = idx + 2
    csv_user = row['User']
    csv_score = row['AnomalyScore']
    csv_breach = row.get('DataBreachEvents', 0)
    csv_alerts = row.get('AlertCount', 0)
    csv_foreign = row.get('ForeignCountrySignIns', 0)

    xl_user = ws.cell(row=excel_row, column=1).value
    xl_score = ws.cell(row=excel_row, column=6).value
    xl_breach = ws.cell(row=excel_row, column=12).value
    xl_alerts = ws.cell(row=excel_row, column=14).value
    xl_foreign = ws.cell(row=excel_row, column=9).value

    user_match = (csv_user == xl_user)
    score_match = (int(csv_score) == int(xl_score))
    breach_match = (int(csv_breach) == int(xl_breach))
    alerts_match = (int(csv_alerts) == int(xl_alerts))
    foreign_match = (int(csv_foreign) == int(xl_foreign))

    if not all([user_match, score_match, breach_match, alerts_match, foreign_match]):
        mismatches += 1
        print(f"MISMATCH #{mismatches}: {csv_user} vs {xl_user}")
        if not user_match: print(f"  User: CSV={csv_user} vs XL={xl_user}")
        if not score_match: print(f"  Score: CSV={csv_score} vs XL={xl_score}")
        if not breach_match: print(f"  Breach: CSV={csv_breach} vs XL={xl_breach}")
        if not alerts_match: print(f"  Alerts: CSV={csv_alerts} vs XL={xl_alerts}")
        if not foreign_match: print(f"  Foreign: CSV={csv_foreign} vs XL={xl_foreign}")

if mismatches == 0:
    print("ALL 54 USERS MATCH PERFECTLY between CSV and Excel!")
else:
    print(f"\nTotal mismatches: {mismatches}")

# Executive Summary check
print("\n=== EXECUTIVE SUMMARY ===")
ws_sum = wb['Executive Summary']
for row in ws_sum.iter_rows(min_row=6, max_row=10, max_col=3, values_only=True):
    print(f"  {row}")

# CSV totals
print("\n=== CSV BASELINE TOTALS ===")
print(f"  AlertCount sum: {csv_df['AlertCount'].sum()}")
print(f"  DataBreachEvents sum: {csv_df['DataBreachEvents'].sum()}")
print(f"  AnomalyScore mean: {csv_df['AnomalyScore'].mean():.1f}")
print(f"  AnomalyScore max: {csv_df['AnomalyScore'].max()}")

# Check verdict totals match
confirmed = len(csv_df[csv_df['Verdict'].str.contains('CONFIRMED', na=False)])
likely = len(csv_df[csv_df['Verdict'].str.contains('Likely Compromised', na=False)])
suspicious = len(csv_df[csv_df['Verdict'].str.contains('Suspicious', na=False)]) - likely  # exclude "Likely" overlap
safe = len(csv_df[csv_df['Verdict'].str.contains('Likely Safe', na=False)])

print(f"\n=== CSV VERDICT COUNTS ===")
print(f"  CONFIRMED: {confirmed}")
print(f"  Likely Compromised: {likely}")
print(f"  Suspicious (excl Likely): {suspicious}")
print(f"  Likely Safe: {safe}")
print(f"  Accounted: {confirmed + likely + suspicious + safe} / {len(csv_df)}")

# Check Action Plan
ws_ap = wb['Action Plan']
p1_count = sum(1 for row in ws_ap.iter_rows(min_row=2, max_col=5, values_only=True) if row[4] and 'P1' in str(row[4]))
p2_count = sum(1 for row in ws_ap.iter_rows(min_row=2, max_col=5, values_only=True) if row[4] and 'P2' in str(row[4]))
print(f"\n=== ACTION PLAN ===")
print(f"  P1 (CRITICAL): {p1_count}")
print(f"  P2 (HIGH): {p2_count}")
