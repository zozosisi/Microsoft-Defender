"""
Excel Investigation Report Generator
=====================================
Professional Excel export using openpyxl (Claude Excel skill patterns).
Generates a multi-sheet workbook with:
  1. Executive Summary — verdict counts, entity breakdown, key metrics
  2. User Investigation — core columns with conditional formatting
  3. Detailed Metrics — full data for technical deep-dive
  4. Action Plan — remediation actions per user

Usage:
    from excel_report import generate_excel_report
    generate_excel_report(summary_df, output_path)
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter


# ============================================================
# COLOR PALETTE (Professional dark theme, muted tones)
# ============================================================
COLORS = {
    # Header
    "header_bg": "1B2A4A",       # Dark navy
    "header_font": "FFFFFF",     # White
    # Verdict fills
    "confirmed_bg": "FADBD8",    # Soft red
    "confirmed_font": "922B21",  # Dark red
    "likely_bg": "FDEBD0",       # Soft orange
    "likely_font": "7E5109",     # Dark orange
    "suspicious_bg": "FEF9E7",   # Soft yellow
    "suspicious_font": "7D6608", # Dark yellow
    "safe_bg": "D5F5E3",         # Soft green
    "safe_font": "1E8449",       # Dark green
    # Misc
    "title_bg": "0D1B2A",       # Very dark navy
    "title_font": "E0E1DD",     # Off-white
    "subtitle_font": "778DA9",  # Steel blue
    "border": "D5D8DC",         # Light gray
    "alt_row": "F8F9F9",        # Very light gray
    "metric_label": "2C3E50",   # Dark slate
    "accent": "2980B9",         # Blue accent
}

THIN_BORDER = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

NO_BORDER = Border()


# ============================================================
# HELPER: Parse JSON array string to readable text
# ============================================================
def parse_json_list(val):
    """Convert JSON array string to comma-separated text."""
    if pd.isna(val) or val == "" or val == "[]":
        return "—"
    try:
        items = json.loads(val)
        if not items:
            return "—"
        return ", ".join(str(i) for i in items)
    except (json.JSONDecodeError, TypeError):
        return str(val)


def verdict_clean(verdict_str):
    """Strip emoji from verdict for Excel."""
    if pd.isna(verdict_str):
        return ""
    return (verdict_str
            .replace("🚨 ", "").replace("🔴 ", "")
            .replace("🟠 ", "").replace("🟢 ", ""))


def verdict_category(verdict_str):
    """Map verdict to category key."""
    if pd.isna(verdict_str):
        return "safe"
    v = str(verdict_str)
    if "CONFIRMED" in v:
        return "confirmed"
    elif "Likely Compromised" in v:
        return "likely"
    elif "Suspicious" in v:
        return "suspicious"
    return "safe"


def get_verdict_fill(category):
    """Get PatternFill for verdict category."""
    return PatternFill(
        start_color=COLORS[f"{category}_bg"],
        end_color=COLORS[f"{category}_bg"],
        fill_type="solid"
    )


def get_verdict_font(category, bold=False):
    """Get Font for verdict category."""
    return Font(
        name="Calibri", size=10,
        color=COLORS[f"{category}_font"],
        bold=bold
    )


# ============================================================
# STYLES
# ============================================================
def create_styles(wb):
    """Register NamedStyles in the workbook."""
    # Header style
    hdr = NamedStyle(name="hdr")
    hdr.font = Font(name="Calibri", size=10, bold=True, color=COLORS["header_font"])
    hdr.fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr.border = Border(bottom=Side(style="medium", color=COLORS["header_bg"]))
    wb.add_named_style(hdr)

    # Data style
    data = NamedStyle(name="data")
    data.font = Font(name="Calibri", size=10, color="2C3E50")
    data.alignment = Alignment(vertical="center", wrap_text=False)
    data.border = THIN_BORDER
    wb.add_named_style(data)

    # Number style (right-aligned)
    num = NamedStyle(name="num")
    num.font = Font(name="Calibri", size=10, color="2C3E50")
    num.alignment = Alignment(horizontal="right", vertical="center")
    num.border = THIN_BORDER
    num.number_format = '#,##0'
    wb.add_named_style(num)

    # Score style (bold number)
    score = NamedStyle(name="score")
    score.font = Font(name="Calibri", size=10, color="2C3E50", bold=True)
    score.alignment = Alignment(horizontal="right", vertical="center")
    score.border = THIN_BORDER
    score.number_format = '#,##0.0'
    wb.add_named_style(score)

    # Title style
    title = NamedStyle(name="title")
    title.font = Font(name="Calibri", size=16, bold=True, color=COLORS["title_font"])
    title.fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
    title.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(title)

    # Subtitle
    sub = NamedStyle(name="sub")
    sub.font = Font(name="Calibri", size=11, color=COLORS["subtitle_font"])
    sub.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(sub)

    # Metric label (left column in summary)
    mlabel = NamedStyle(name="mlabel")
    mlabel.font = Font(name="Calibri", size=11, bold=True, color=COLORS["metric_label"])
    mlabel.alignment = Alignment(horizontal="left", vertical="center")
    mlabel.border = Border(bottom=Side(style="thin", color=COLORS["border"]))
    wb.add_named_style(mlabel)

    # Metric value
    mval = NamedStyle(name="mval")
    mval.font = Font(name="Calibri", size=11, color=COLORS["metric_label"])
    mval.alignment = Alignment(horizontal="left", vertical="center")
    mval.border = Border(bottom=Side(style="thin", color=COLORS["border"]))
    wb.add_named_style(mval)


def auto_width(ws, min_width=8, max_width=45):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def write_header_row(ws, row, headers, style="hdr"):
    """Write a header row with style."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.style = style


def apply_verdict_formatting(ws, row, verdict_str, col_start=1, col_end=None):
    """Apply verdict color to an entire data row."""
    cat = verdict_category(verdict_str)
    fill = get_verdict_fill(cat)
    font = get_verdict_font(cat)
    if col_end is None:
        col_end = ws.max_column
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER


# ============================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================
def build_executive_summary(ws, df, threshold):
    """Build the Executive Summary sheet."""
    ws.sheet_properties.tabColor = COLORS["accent"]

    # Title block (merged)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Crystal Group — Security Investigation Report"
    title_cell.style = "title"
    ws.row_dimensions[1].height = 40

    # Subtitle row
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Users: {len(df)}  |  Threshold: {threshold*100:.0f}%"
    ws["A2"].style = "sub"
    ws.row_dimensions[2].height = 22

    # --- Verdict Summary ---
    row = 4
    ws.cell(row=row, column=1, value="VERDICT SUMMARY").style = "mlabel"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    verdict_types = [
        ("CONFIRMED COMPROMISED", "confirmed"),
        ("Likely Compromised", "likely"),
        ("Suspicious", "suspicious"),
        ("Likely Safe", "safe"),
    ]

    headers = ["Verdict", "Users", "% of Total"]
    write_header_row(ws, row, headers)
    row += 1

    # Display labels for each verdict (clean, no emoji)
    verdict_labels = {
        "confirmed": "CONFIRMED COMPROMISED (Data Breach)",
        "likely": "Likely Compromised",
        "suspicious": "Suspicious",
        "safe": "Likely Safe",
    }

    for vtext, cat in verdict_types:
        count = len(df[df["Verdict"].str.contains(vtext, na=False, regex=False)])
        pct = count / len(df) * 100 if len(df) > 0 else 0

        cell_v = ws.cell(row=row, column=1, value=verdict_labels[cat])
        cell_c = ws.cell(row=row, column=2, value=count)
        cell_p = ws.cell(row=row, column=3, value=f"{pct:.1f}%")

        fill = get_verdict_fill(cat)
        font = get_verdict_font(cat, bold=True)
        for c in [cell_v, cell_c, cell_p]:
            c.fill = fill
            c.font = font
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
        cell_c.alignment = Alignment(horizontal="center", vertical="center")
        cell_p.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # Total row
    total_font = Font(name="Calibri", size=10, bold=True, color=COLORS["header_font"])
    total_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    for c in range(1, 4):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["header_font"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        cell.border = THIN_BORDER
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=2).value = len(df)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # --- Entity Breakdown ---
    ws.cell(row=row, column=1, value="ENTITY BREAKDOWN").style = "mlabel"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    write_header_row(ws, row, ["Entity", "Users", "Compromised"])
    row += 1

    for entity in ["ABL", "CMBD", "CETBD", "OTHER"]:
        ent_df = df[df["Entity"] == entity]
        compromised = len(ent_df[ent_df["Verdict"].str.contains("Compromised", na=False, regex=False)])
        ws.cell(row=row, column=1, value=entity).style = "data"
        ws.cell(row=row, column=2, value=len(ent_df)).style = "num"
        ws.cell(row=row, column=3, value=compromised).style = "num"
        row += 1
    row += 1

    # --- Key Metrics ---
    ws.cell(row=row, column=1, value="KEY METRICS").style = "mlabel"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    metrics = [
        ("Total Alerts (all users)", df["AlertCount"].sum() if "AlertCount" in df.columns else 0),
        ("Total Data Breach Events", df["DataBreachEvents"].sum() if "DataBreachEvents" in df.columns else 0),
        ("Users with Phishing Emails", len(df[df.get("PhishingEmailsReceived", 0) > 0]) if "PhishingEmailsReceived" in df.columns else 0),
        ("Users with Admin Role", len(df[df["IsAdmin"] == True]) if "IsAdmin" in df.columns else 0),
        ("Avg Anomaly Score", round(df["AnomalyScore"].mean(), 1) if "AnomalyScore" in df.columns else 0),
        ("Max Anomaly Score", df["AnomalyScore"].max() if "AnomalyScore" in df.columns else 0),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label).style = "mlabel"
        ws.cell(row=row, column=2, value=value).style = "mval"
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


# ============================================================
# SHEET 2: USER INVESTIGATION (core columns)
# ============================================================
def build_user_investigation(ws, df):
    """Build the User Investigation sheet with core columns."""
    ws.sheet_properties.tabColor = "2980B9"

    columns = [
        ("User (UPN)", "User", "data", 35),
        ("Display Name", "DisplayName", "data", 30),
        ("Entity", "Entity", "data", 8),
        ("Department", "Department", "data", 25),
        ("Verdict", None, "data", 38),   # special handling
        ("Anomaly Score", "AnomalyScore", "score", 14),
        ("Total Sign-ins", "TotalSignIns", "num", 14),
        ("Active Days (30d)", "ActiveDays", "num", 14),
        ("Foreign Sign-ins (count)", "ForeignCountrySignIns", "num", 20),
        ("Foreign Countries (list)", "ForeignCountryList", "data", 50),
        ("Suspicious IPs (unique)", None, "num", 18),  # derived
        ("Data Breach Events", "DataBreachEvents", "num", 16),
        ("Data Breach Actions", "DataBreachActions", "data", 30),
        ("Defender Alert Count", "AlertCount", "num", 16),
        ("MFA Status", "MFAStatus", "data", 32),
        ("Account Status", "AccountStatus", "data", 14),
        ("Is Admin", "IsAdmin", "data", 10),
        ("Baseline Warning", "BaselineWarning", "data", 50),
    ]

    # Header row
    for col_idx, (label, _, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.style = "hdr"
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        verdict_raw = row_data.get("Verdict", "")

        for col_idx, (label, field, style, _) in enumerate(columns, 1):
            # Determine value
            if label == "Verdict":
                value = verdict_clean(verdict_raw)
            elif label == "Suspicious IPs (unique)":
                try:
                    value = len(json.loads(row_data.get("SuspiciousIPs", "[]")))
                except (json.JSONDecodeError, TypeError):
                    value = 0
            elif label == "Foreign Countries (list)":
                value = parse_json_list(row_data.get("ForeignCountryList", "[]"))
            elif label == "Data Breach Actions":
                value = parse_json_list(row_data.get("DataBreachActions", "[]"))
            elif label == "Is Admin":
                value = "YES" if row_data.get("IsAdmin", False) else "No"
            elif field:
                value = row_data.get(field, "")
                if pd.isna(value):
                    value = ""
            else:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.style = style

        # Apply verdict row coloring
        apply_verdict_formatting(ws, row_idx, verdict_raw, 1, len(columns))
        # Keep score bold
        score_cell = ws.cell(row=row_idx, column=6)
        cat = verdict_category(verdict_raw)
        score_cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS[f"{cat}_font"])

    # Freeze panes & auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(df)+1}"


# ============================================================
# SHEET 3: DETAILED METRICS (all columns)
# ============================================================
def build_detailed_metrics(ws, df):
    """Build the Detailed Metrics sheet with all columns."""
    ws.sheet_properties.tabColor = "7F8C8D"

    # JSON columns that need parsing
    json_cols = {
        "TrustedIPs", "TrustedCountries", "AllCountries", "TrustedCities",
        "TrustedDevices", "TrustedBrowsers", "TrustedOS", "ISPList",
        "UnknownIPList", "ForeignCountryList", "NonBDCountries",
        "SuspiciousISPs", "SuspiciousIPs", "DataBreachActions",
        "AlertIPsMatched", "CrossUserAlertIPs",
    }

    headers = list(df.columns)

    # Write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.style = "hdr"

    # Write data
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        verdict_raw = row_data.get("Verdict", "")
        for col_idx, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            if pd.isna(val):
                val = ""
            # Parse JSON columns
            if h in json_cols and isinstance(val, str):
                val = parse_json_list(val)
            # Clean verdict emoji
            if h == "Verdict":
                val = verdict_clean(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Use num style for numeric columns
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.style = "num"
            else:
                cell.style = "data"

        apply_verdict_formatting(ws, row_idx, verdict_raw, 1, len(headers))

    # Freeze panes & auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df)+1}"
    auto_width(ws, max_width=35)


# ============================================================
# SHEET 4: ACTION PLAN
# ============================================================
def build_action_plan(ws, df):
    """Build the Action Plan sheet with remediation per user."""
    ws.sheet_properties.tabColor = "E74C3C"

    columns = [
        ("User", 35),
        ("Display Name", 28),
        ("Entity", 8),
        ("Verdict", 30),
        ("Priority", 10),
        ("Immediate Actions", 75),
        ("Status", 12),
    ]

    # Headers
    for col_idx, (label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.style = "hdr"
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Action templates
    actions = {
        "confirmed": (
            "P1 — CRITICAL",
            "1. Revoke all active sessions immediately\n"
            "2. Reset password\n"
            "3. Re-enroll MFA (FIDO2 preferred)\n"
            "4. Review & remove suspicious mailbox rules\n"
            "5. Forensic review of breached files (FileAccessed/FileDownloaded)\n"
            "6. Notify affected data owners"
        ),
        "likely": (
            "P2 — HIGH",
            "1. Revoke all active sessions\n"
            "2. Reset password\n"
            "3. Monitor sign-in activity for 72 hours\n"
            "4. Review mailbox rules for suspicious forwarding"
        ),
        "suspicious": (
            "P3 — MEDIUM",
            "1. Monitor sign-in activity\n"
            "2. Verify with user if activity is legitimate\n"
            "3. Consider password reset if unverified"
        ),
        "safe": (
            "P4 — LOW",
            "No immediate action required. Continue monitoring."
        ),
    }

    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        verdict_raw = row_data.get("Verdict", "")
        cat = verdict_category(verdict_raw)
        priority, action_text = actions[cat]

        values = [
            row_data.get("User", ""),
            row_data.get("DisplayName", ""),
            row_data.get("Entity", ""),
            verdict_clean(verdict_raw),
            priority,
            action_text,
            "Pending",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.style = "data"
            if col_idx == 6:  # Actions column — wrap text
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        apply_verdict_formatting(ws, row_idx, verdict_raw, 1, len(columns))

        # Priority cell bold
        p_cell = ws.cell(row=row_idx, column=5)
        p_cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS[f"{cat}_font"])

        # Row height for action text
        ws.row_dimensions[row_idx].height = 75 if cat in ("confirmed", "likely") else 40

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(df)+1}"


# ============================================================
# SHEET 5: SCORING LOGIC (verdict methodology)
# ============================================================
def build_scoring_logic(ws):
    """Build the Scoring Logic sheet documenting verdict methodology."""
    ws.sheet_properties.tabColor = "8E44AD"  # Purple

    # --- Title ---
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Verdict Scoring Logic — Detection Methodology v5.0"
    title_cell.style = "title"
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"].value = "This sheet documents how each user's Anomaly Score and Verdict are calculated."
    ws["A2"].style = "sub"
    ws.row_dimensions[2].height = 22

    # --- Section 1: Scoring Matrix ---
    row = 4
    ws.cell(row=row, column=1, value="SCORING MATRIX").style = "mlabel"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    headers = ["Risk Factor", "Points", "Max Cap", "Condition", "Data Source", "Description"]
    write_header_row(ws, row, headers)
    row += 1

    scoring_rules = [
        ("Data Breach Actions", "+1,000", "—",
         "DataBreachEvents > 0", "Q09 CloudAppEvents",
         "File accessed/downloaded from suspicious IP. Instant CONFIRMED verdict."),
        ("Hacker Botnet Countries", "+30 / country", "No cap",
         "Country in sign-in but NOT in TrustedCountries", "Q01 Sign-in History",
         "Foreign countries from unknown devices. Dedup: countries with VPN sign-ins excluded."),
        ("Suspicious IPs", "+5 / IP", "Max 50",
         "Unknown IP + Unknown Device", "Q01 Sign-in History",
         "IPs never seen with a trusted device. Catches domestic hackers using local ISP."),
        ("Suspicious ISPs", "+15 / ISP", "No cap",
         "ISP is Hosting/VPS/Anonymous", "Q02 ISP Data",
         "Sign-ins from hosting providers, VPS, or anonymizing networks."),
        ("AiTM Token Theft", "+15 / session", "Max 45",
         "Multi-IP session + MFA-by-token + Unknown Device", "Q01 AuthProcessingDetails",
         "Session cookie stolen via AiTM proxy. Only scored when MFA bypass by token detected."),
        ("Benign Unknown IPs", "+2 / IP", "Max 30",
         "Unknown IP + Trusted Device", "Q01 Sign-in History",
         "Likely travel/VPN — same device, new IP. Low penalty."),
        ("Entra ID Risk Events", "+5 / event", "No cap",
         "RiskLevel >= 50 (Medium/High)", "Q01 Sign-in History",
         "Microsoft flagged the session as risky (atypical travel, leaked credentials, etc.)"),
        ("Defender Alerts", "+5 / alert", "Max 25",
         "CONDITIONAL: only if user has compromise indicators", "Q03 Alert Data",
         "Alerts only count when user also has foreign sign-ins, suspicious IPs, or high-risk events."),
        ("Alert IP Correlation", "+3 / IP match", "Max 30",
         "Suspicious IP also appears in Q00 Alert IPs", "Q00 Incidents",
         "Suspicious IP confirmed by Entra ID Protection as unfamiliar sign-in source."),
        ("Phishing Target", "+5 / email", "No cap",
         "User received phishing email", "Q05 Phishing Check",
         "User was targeted by phishing campaign — increases likelihood of credential theft."),
        ("Off-Hours Sign-ins", "+0.5 / event", "Max 10",
         "Sign-in outside business hours", "Q01 Sign-in History",
         "Sign-ins at unusual hours. Low weight — many legitimate cases."),
        ("Unmanaged Devices", "+5 (flat)", "Max 5",
         "Unmanaged device % > 80%", "Q01 Sign-in History",
         "User primarily uses personal/unmanaged devices."),
        ("Admin Account Boost", "+10 (flat)", "Max 10",
         "IsAdmin = true AND score >= 15", "Q10 Auth Status",
         "Extra penalty for admin accounts already showing suspicious activity."),
        ("VPN Countries", "+0", "—",
         "Country in sign-in + Trusted Device", "Q01 Sign-in History",
         "Legitimate VPN usage — foreign country but on known/trusted device. No penalty."),
    ]

    for rule in scoring_rules:
        for col_idx, val in enumerate(rule, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.style = "data"
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        # Highlight Data Breach row
        if "1,000" in rule[1]:
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.fill = PatternFill(start_color=COLORS["confirmed_bg"], end_color=COLORS["confirmed_bg"], fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color=COLORS["confirmed_font"])
        # Highlight VPN row (safe — green)
        if rule[1] == "+0":
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.fill = PatternFill(start_color=COLORS["safe_bg"], end_color=COLORS["safe_bg"], fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color=COLORS["safe_font"])
        ws.row_dimensions[row].height = 36
        row += 1

    # --- Section 2: Verdict Thresholds ---
    row += 1
    ws.cell(row=row, column=1, value="VERDICT CLASSIFICATION").style = "mlabel"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    write_header_row(ws, row, ["Verdict", "Score Threshold", "Meaning", "", "", ""])
    row += 1

    verdict_rules = [
        ("CONFIRMED COMPROMISED (Data Breach)", "DataBreachEvents > 0",
         "Hacker accessed/downloaded files from suspicious IP. Immediate remediation required.",
         "confirmed"),
        ("Likely Compromised", "Score >= 30",
         "Multiple strong indicators of account compromise. Password reset + session revoke recommended.",
         "likely"),
        ("Suspicious", "Score >= 15",
         "Some anomalous activity detected. Monitor and verify with user.",
         "suspicious"),
        ("Likely Safe", "Score < 15",
         "No significant anomalies. Alerts may be false positives from VPN/travel.",
         "safe"),
        ("Verified Safe (SOC Override)", "Manual override",
         "SOC team has verified the user's activity is legitimate. Score forced to 0.",
         "safe"),
    ]

    for verdict, threshold, meaning, cat in verdict_rules:
        fill = get_verdict_fill(cat)
        font = get_verdict_font(cat)
        for col_idx, val in enumerate([verdict, threshold, meaning], 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.font = font
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if col_idx == 3:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 32
        row += 1

    # --- Section 3: Data Sources ---
    row += 1
    ws.cell(row=row, column=1, value="DATA SOURCES (KQL Queries)").style = "mlabel"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    write_header_row(ws, row, ["Query", "Export File", "Table", "Purpose", "", ""])
    row += 1

    sources = [
        ("Q00", "unfamiliar_signin_incidents.csv", "AlertInfo + AlertEvidence",
         "Master query — all Unfamiliar Sign-in incidents, users, IPs"),
        ("Q01A-F", "signin_history_01..06.csv -> merged", "EntraIdSignInEvents",
         "Sign-in history (30 days, split 6 parts) + AuthProcessingDetails for AiTM"),
        ("Q02", "isp_data.csv", "IdentityLogonEvents",
         "ISP enrichment — identify hosting/VPS/anonymous ISPs"),
        ("Q03", "alert_data.csv", "AlertEvidence",
         "Unfamiliar sign-in alert details per user"),
        ("Q04", "user_profiles.csv", "IdentityInfo",
         "User identity info — department, job title, risk level"),
        ("Q05", "phishing_emails.csv", "EmailEvents",
         "Phishing emails received by affected users"),
        ("Q09", "cloudapp_events.csv", "CloudAppEvents",
         "Cloud app activity — file access/download from suspicious IPs (data breach)"),
        ("Q10", "auth_status.csv", "IdentityAccountInfo",
         "MFA status, password reset history, admin roles"),
    ]

    for query, file, table, purpose in sources:
        for col_idx, val in enumerate([query, file, table, purpose], 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.style = "data"
        if col_idx == 4:
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        row += 1

    # --- Section 4: Key Concepts ---
    row += 1
    ws.cell(row=row, column=1, value="KEY CONCEPTS").style = "mlabel"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    concepts = [
        ("Trusted IP", "IP appearing in >= 5% of user's total sign-ins (dynamic per-user baseline)"),
        ("Trusted Country", "Country appearing in >= 5% of sign-ins (15% threshold for low-volume users < 50 sign-ins)"),
        ("Trusted Device", "Device used >= 5% of sign-ins — known corporate/personal device"),
        ("VPN Country", "Foreign country sign-in on a Trusted Device = legitimate VPN usage, not penalized"),
        ("Hacker Botnet Country", "Foreign country sign-in on Unknown Device = likely residential proxy/botnet"),
        ("Suspicious IP", "Unknown IP + Unknown Device sign-in = high-confidence attacker indicator"),
        ("Benign Unknown IP", "Unknown IP + Trusted Device = likely travel or new VPN endpoint"),
        ("MS Infra IPs", "Microsoft infrastructure IPs (20.x, 40.x, 52.x, etc.) are auto-filtered before analysis"),
        ("AiTM Session", "Same SessionId from different IPs. Multi-device login (PC + Phone) creates separate SessionIds — NOT affected. Only scored when combined with MFA-by-token bypass + Unknown Device."),
        ("Baseline Contamination", "If > 15 Trusted Countries detected, attacker may have poisoned the baseline"),
    ]

    for term, definition in concepts:
        cell_t = ws.cell(row=row, column=1, value=term)
        cell_t.style = "mlabel"
        cell_d = ws.cell(row=row, column=2, value=definition)
        cell_d.style = "mval"
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 45


# ============================================================
# MAIN ENTRY POINT
# ============================================================
def generate_excel_report(df: pd.DataFrame, output_path: Path, threshold: float = 0.05):
    """Generate a professional multi-sheet Excel investigation report.

    Args:
        df: Summary DataFrame (sorted by AnomalyScore desc)
        output_path: Path to save the .xlsx file
        threshold: Trusted threshold used in analysis
    """
    wb = Workbook()
    create_styles(wb)

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    build_executive_summary(ws1, df, threshold)

    # Sheet 2: User Investigation
    ws2 = wb.create_sheet("User Investigation")
    build_user_investigation(ws2, df)

    # Sheet 3: Detailed Metrics
    ws3 = wb.create_sheet("Detailed Metrics")
    build_detailed_metrics(ws3, df)

    # Sheet 4: Action Plan
    ws4 = wb.create_sheet("Action Plan")
    build_action_plan(ws4, df)

    # Sheet 5: Scoring Logic
    ws5 = wb.create_sheet("Scoring Logic")
    build_scoring_logic(ws5)

    # Save
    wb.save(output_path)
    print(f"\U0001f4ca Excel report saved: {output_path}")

